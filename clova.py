# goolge colab 에서 작성

from openpyxl import Workbook
from openpyxl import load_workbook
from konlpy.tag import Okt
import re
import sys
import requests
import json


url = "https://naveropenapi.apigw.ntruss.com/sentiment-analysis/v1/analyze"

headers = {   # 지원 api key
    "X-NCP-APIGW-API-KEY-ID": "x4cybkvtso",
    "X-NCP-APIGW-API-KEY": "QX0KLAU5yVSuSayqr1t285G9hvI3YEWq09fIVQpv",
    "Content-Type": "application/json"
}


load_wb = load_workbook(
    "/content/drive/MyDrive/jw/seoul/yongsan_restaurants.xlsx", data_only=True)

load_ws = load_wb['용산구']
wb = Workbook()
ws = wb.active

review_t = []
review = []

for row in load_ws.rows:
    row_value = []
    for cell in row:
        if cell.value != None:
            row_value.append(cell.value)
    review_t.append(row_value)

for i in review_t:
    t = ''
    for j in i:
        t = t+" "+j
    review.append(t)

# print(review)

okt = Okt()
word = []

for i in range(len(review)):
    review[i] = re.sub("[^가-힣ㄱ-ㅎㅏ-ㅣ\\s]", "", review[i])  # 한글이랑 공백만 남기고 다 지움
    review[i] = review[i].replace("\n", " ")
    word.append(review[i])

# print(word)
print("매장 개수", len(word))  # len(review) == len(word)  매장 개수

clova_cnt = 0
for i in range(len(word)):
    if (int(len(word[i])/950) == 0):
        clova_cnt += 1
    else:
        clova_cnt += int(len(word[i])/950)

print("api 호출 횟수", clova_cnt)  # api 호출 횟수
print("\n")

p_li = []
n_li = []

for i in range(len(word)):
    pp = 0
    nn = 0
    pp_avg = 0
    nn_avg = 0
    n = 0

    if(len(word[i]) > 950):
        length = int(len(word[i])/950)   # 매장 하나의 전체 리뷰 글자수를 950으로 나눈 값만큼 돌리기
        m = 950
    else:
        length = 1                       # 950 미만이면 한번만
        m = len(word[i])

    for k in range(length):

        content = word[i][n:m]

        data = {
            "content": content
        }

        response = requests.post(url, data=json.dumps(data), headers=headers)
        rescode = response.status_code

        text = response.json()

        documnet = text['document']
        sentiment = text['document']['sentiment']
        positive = text['document']['confidence']['positive']
        negative = text['document']['confidence']['negative']

        pp += positive
        nn += negative

        n += 950
        m += 950

        pp_avg = pp/length
        nn_avg = nn/length

    if(rescode == 200):
        print(review_t[i][0], pp_avg, nn_avg)
        p_li.append([pp_avg, review_t[i][0]])
        n_li.append([nn_avg, review_t[i][0]])
    else:
        print("Error : " + response.text)

p_li.sort(reverse=True)
n_li.sort(reverse=True)

p_result = 0
n_result = 0

print("\n긍정")
for i in range(len(word)):
    p_result = "{:,.1f}".format(p_li[i][0])
    print(p_li[i][1], p_result)

print("\n부정")
for i in range(len(word)):
    n_result = "{:,.1f}".format(n_li[i][0])
    print(n_li[i][1], n_result)
